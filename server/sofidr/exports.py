"""Portable analysis report exports for SOFIDR results."""

from __future__ import annotations

import html
import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


REPORT_FORMATS = {"json", "html", "pdf", "xlsx"}


def _title(value: str) -> str:
    return value.replace("_", " ").title()


def _scoreboard(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list(payload.get("scoreboard") or [])


def _json_report(payload: dict[str, Any]) -> bytes:
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


def _html_report(payload: dict[str, Any]) -> bytes:
    rows = "".join(
        "<tr>"
        f"<td>{index}</td><td>{html.escape(_title(str(item['name'])))}</td>"
        f"<td>{float(item['sei']):.4f}</td><td>{float(item['accuracy']):.4f}</td>"
        f"<td>{float(item['stability']):.4f}</td><td>{float(item['retention']):.3f}</td>"
        f"<td>{float(item['simplicity']):.3f}</td>"
        "</tr>"
        for index, item in enumerate(_scoreboard(payload), 1)
    )
    tags = " · ".join(html.escape(str(tag)) for tag in payload.get("terrain_tags", []))
    document = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>SOFIDR report — {html.escape(str(payload['dataset_name']))}</title>
  <style>
    body {{ color:#17211b; font:15px/1.5 Arial,sans-serif; margin:40px auto; max-width:1080px; }}
    h1,h2 {{ color:#0b3d2e; }} .meta {{ color:#526057; }}
    .recommendation {{ background:#edf7f1; border-left:5px solid #1b7f5a; padding:18px; }}
    table {{ border-collapse:collapse; margin:24px 0; width:100%; }}
    th,td {{ border:1px solid #ccd7d0; padding:8px; text-align:right; }}
    th:nth-child(2),td:nth-child(2) {{ text-align:left; }} th {{ background:#0b3d2e; color:white; }}
    pre {{ background:#f4f6f4; padding:18px; white-space:pre-wrap; }}
  </style>
</head>
<body>
  <p class="meta">SOFIDR · Strategic Optimization Framework for Iterative Data Regeneration</p>
  <h1>{html.escape(str(payload['dataset_name']))}</h1>
  <p class="meta">{tags}</p>
  <section class="recommendation">
    <strong>Best strategic fit</strong>
    <h2>{html.escape(_title(str(payload['best_by_sei'])))}</h2>
    <p>Learning policy: {html.escape(_title(str(payload['selected'])))}</p>
    <p>{html.escape(str(payload['selection_reason']))}</p>
  </section>
  <h2>Formation scoreboard</h2>
  <table>
    <thead><tr><th>Rank</th><th>Formation</th><th>SEI</th><th>Accuracy</th><th>Stability</th><th>Retention</th><th>Simplicity</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
  <h2>Decision record</h2>
  <pre>{html.escape(str(payload['report']))}</pre>
</body>
</html>"""
    return document.encode("utf-8")


def _pdf_report(payload: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
        title=f"SOFIDR report — {payload['dataset_name']}",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("SOFIDR analysis report", styles["Title"]),
        Paragraph(html.escape(str(payload["dataset_name"])), styles["Heading2"]),
        Paragraph(
            f"Best strategic fit: <b>{html.escape(_title(str(payload['best_by_sei'])))}</b>",
            styles["BodyText"],
        ),
        Paragraph(
            "Terrain: "
            + html.escape(" · ".join(str(tag) for tag in payload.get("terrain_tags", []))),
            styles["BodyText"],
        ),
        Spacer(1, 0.18 * inch),
    ]
    table_data = [["Rank", "Formation", "SEI", "Accuracy", "Stability", "Retention", "Simplicity"]]
    for index, item in enumerate(_scoreboard(payload), 1):
        table_data.append(
            [
                index,
                _title(str(item["name"])),
                f"{float(item['sei']):.4f}",
                f"{float(item['accuracy']):.4f}",
                f"{float(item['stability']):.4f}",
                f"{float(item['retention']):.3f}",
                f"{float(item['simplicity']):.3f}",
            ]
        )
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[0.5 * inch, 1.45 * inch, 0.8 * inch, 0.9 * inch, 0.9 * inch, 0.9 * inch, 0.9 * inch],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b3d2e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#ccd7d0")),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f4")]),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend(
        [
            table,
            Spacer(1, 0.22 * inch),
            Paragraph("Decision record", styles["Heading2"]),
            Paragraph(
                html.escape(str(payload["report"])).replace("\n", "<br/>"),
                styles["BodyText"],
            ),
        ]
    )
    document.build(story)
    return buffer.getvalue()


def _xlsx_report(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["SOFIDR analysis report"])
    summary.append(["Dataset", payload["dataset_name"]])
    summary.append(["Best formation", _title(str(payload["best_by_sei"]))])
    summary.append(["Learning policy", _title(str(payload["selected"]))])
    summary.append(["Selection reason", payload["selection_reason"]])
    summary.append(["Terrain", ", ".join(str(tag) for tag in payload.get("terrain_tags", []))])
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="0B3D2E")
    summary.merge_cells("A1:B1")
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 90
    summary["B5"].alignment = Alignment(wrap_text=True, vertical="top")

    scores = workbook.create_sheet("Scoreboard")
    headers = ["Rank", "Formation", "SEI", "Accuracy", "Stability", "Retention", "Simplicity", "Error"]
    scores.append(headers)
    for index, item in enumerate(_scoreboard(payload), 1):
        scores.append(
            [
                index,
                _title(str(item["name"])),
                item["sei"],
                item["accuracy"],
                item["stability"],
                item["retention"],
                item["simplicity"],
                item.get("error", ""),
            ]
        )
    for cell in scores[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B3D2E")
    scores.freeze_panes = "A2"
    scores.auto_filter.ref = scores.dimensions
    for column, width in {"A": 9, "B": 24, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 30}.items():
        scores.column_dimensions[column].width = width

    decision = workbook.create_sheet("Decision Record")
    decision["A1"] = "Decision record"
    decision["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    decision["A1"].fill = PatternFill("solid", fgColor="0B3D2E")
    decision["A2"] = payload["report"]
    decision["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    decision.column_dimensions["A"].width = 120
    decision.row_dimensions[2].height = 300

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_report(payload: dict[str, Any], report_format: str) -> tuple[bytes, str, str]:
    """Return report bytes, media type, and extension."""
    if report_format not in REPORT_FORMATS:
        raise ValueError(f"Unsupported report format: {report_format}")
    builders = {
        "json": (_json_report, "application/json", "json"),
        "html": (_html_report, "text/html", "html"),
        "pdf": (_pdf_report, "application/pdf", "pdf"),
        "xlsx": (
            _xlsx_report,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        ),
    }
    builder, media_type, extension = builders[report_format]
    return builder(payload), media_type, extension
