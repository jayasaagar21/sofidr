import io
import json

import pandas as pd
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from main import app


client = TestClient(app)


def _csv(rows: int = 18, labels=(0, 1)) -> bytes:
    frame = pd.DataFrame(
        {
            "feature_a": [float(i) for i in range(rows)],
            "feature_b": [float((i * 7) % 11) for i in range(rows)],
            "target": [labels[i % len(labels)] for i in range(rows)],
        }
    )
    return frame.to_csv(index=False).encode()


def test_health_and_security_headers():
    response = client.get("/api/health")
    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": "1.0.0"}
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["x-frame-options"] == "DENY"


def test_metadata_endpoints_are_populated():
    archetypes = client.get("/api/archetypes")
    formations = client.get("/api/formations")
    assert archetypes.status_code == 200
    assert "breast_cancer" in archetypes.json()["archetypes"]
    assert formations.status_code == 200
    assert {"phalanx", "guerrilla", "reconnaissance"} <= set(formations.json())


def test_upload_accepts_negative_class_labels():
    response = client.post(
        "/api/optimize?model=logreg",
        files={"file": ("negative-labels.csv", io.BytesIO(_csv(labels=(-1, 1))), "text/csv")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["success"] is True
    assert body["best_by_sei"]
    assert len(body["scoreboard"]) == 8


def test_upload_rejects_continuous_target():
    frame = pd.DataFrame(
        {
            "feature": range(20),
            "target": [i / 10 for i in range(20)],
        }
    )
    response = client.post(
        "/api/optimize",
        files={
            "file": (
                "regression.csv",
                io.BytesIO(frame.to_csv(index=False).encode()),
                "text/csv",
            )
        },
    )
    assert response.status_code == 400
    assert "discrete classes" in response.json()["detail"]


def test_upload_rejects_infinite_features():
    payload = b"feature,target\n1,0\n2,0\n3,0\n4,1\n5,1\ninf,1\n"
    response = client.post(
        "/api/optimize",
        files={"file": ("infinite.csv", io.BytesIO(payload), "text/csv")},
    )
    assert response.status_code == 400
    assert "infinite values" in response.json()["detail"]


def test_missing_input_is_a_client_error():
    response = client.post("/api/optimize")
    assert response.status_code == 400
    assert "Provide either" in response.json()["detail"]


def test_enhance_returns_parseable_attachment_and_metadata():
    frame = pd.DataFrame(
        {
            "feature_a": [float(i) for i in range(20)],
            "label": ["majority"] * 15 + ["minority"] * 5,
            "feature_b": [float((i * 3) % 7) for i in range(20)],
        }
    )
    response = client.post(
        "/api/enhance?formation=guerrilla&target_column=label",
        files={
            "file": (
                "source data.csv",
                io.BytesIO(frame.to_csv(index=False).encode()),
                "text/csv",
            )
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert response.headers["content-disposition"] == (
        'attachment; filename="source_data-guerrilla-sofidr.csv"'
    )
    assert response.headers["x-sofidr-input-rows"] == "20"
    assert response.headers["x-sofidr-output-rows"] == "30"
    assert response.headers["x-sofidr-synthetic-rows"] == "10"
    assert response.headers["x-sofidr-removed-rows"] == "0"
    assert response.headers["x-sofidr-steps"] == "impute,smote,scale"

    enhanced = pd.read_csv(io.BytesIO(response.content))
    assert enhanced.columns.tolist() == [
        "feature_a", "feature_b", "label", "_sofidr_row_origin"
    ]
    assert enhanced["label"].value_counts().to_dict() == {
        "majority": 15,
        "minority": 15,
    }


def test_enhance_uses_all_rows_beyond_optimization_sample():
    response = client.post(
        "/api/enhance?formation=reconnaissance",
        files={"file": ("large.csv", io.BytesIO(_csv(rows=350)), "text/csv")},
    )

    assert response.status_code == 200
    assert response.headers["x-sofidr-input-rows"] == "350"
    assert response.headers["x-sofidr-output-rows"] == "350"
    assert len(pd.read_csv(io.BytesIO(response.content))) == 350


def test_enhance_rejects_unknown_formation_and_malformed_csv():
    unknown = client.post(
        "/api/enhance?formation=unknown",
        files={"file": ("data.csv", io.BytesIO(_csv()), "text/csv")},
    )
    malformed = client.post(
        "/api/enhance?formation=phalanx",
        files={"file": ("bad.csv", io.BytesIO(b'a,b\n1,"unterminated'), "text/csv")},
    )

    assert unknown.status_code == 400
    assert "Unknown formation" in unknown.json()["detail"]
    assert malformed.status_code == 400
    assert "Invalid CSV" in malformed.json()["detail"]


def test_enhance_requires_enough_rows_per_class_for_smote():
    frame = pd.DataFrame(
        {
            "feature": range(10),
            "target": [0] * 9 + [1],
        }
    )
    response = client.post(
        "/api/enhance?formation=guerrilla",
        files={"file": ("tiny-class.csv", io.BytesIO(frame.to_csv(index=False).encode()), "text/csv")},
    )

    assert response.status_code == 400
    assert "at least 2 rows" in response.json()["detail"]


def _report_payload():
    return {
        "success": True,
        "dataset_name": "customer <risk>.csv",
        "terrain_tags": ["balanced", "moderate_dimensionality"],
        "cold_start_default": "reconnaissance",
        "best_by_sei": "phalanx",
        "selected": "phalanx",
        "selection_reason": "Highest deterministic SEI score.",
        "scoreboard": [
            {
                "name": "phalanx",
                "sei": 0.91,
                "accuracy": 0.92,
                "stability": 0.89,
                "retention": 0.95,
                "simplicity": 0.8,
                "error": "",
            },
            {
                "name": "reconnaissance",
                "sei": 0.86,
                "accuracy": 0.88,
                "stability": 0.84,
                "retention": 1.0,
                "simplicity": 1.0,
                "error": "",
            },
        ],
        "report": "SOFIDR decision record\nBest formation: phalanx",
        "error": "",
    }


def test_report_exports_json_html_pdf_and_excel():
    responses = {
        report_format: client.post(
            f"/api/report?format={report_format}",
            json=_report_payload(),
        )
        for report_format in ("json", "html", "pdf", "xlsx")
    }

    for report_format, response in responses.items():
        assert response.status_code == 200, (report_format, response.text)
        assert response.headers["content-disposition"].endswith(
            f'-sofidr-report.{report_format}"'
        )
        assert response.headers["x-sofidr-report-format"] == report_format

    assert json.loads(responses["json"].content)["best_by_sei"] == "phalanx"
    assert b"customer &lt;risk&gt;.csv" in responses["html"].content
    assert responses["pdf"].content.startswith(b"%PDF")

    workbook = load_workbook(io.BytesIO(responses["xlsx"].content), read_only=True)
    assert workbook.sheetnames == ["Summary", "Scoreboard", "Decision Record"]
    assert workbook["Summary"]["B3"].value == "Phalanx"


def test_report_rejects_failed_analysis_and_unknown_format():
    failed = _report_payload()
    failed["success"] = False
    failed["error"] = "analysis failed"

    failed_response = client.post("/api/report?format=pdf", json=failed)
    unsupported_response = client.post("/api/report?format=docx", json=_report_payload())

    assert failed_response.status_code == 400
    assert unsupported_response.status_code == 422


def test_clean_accepts_malformed_mixed_business_dataset():
    payload = (
        "Transaction_ID,Customer_Name,Email,Age,Purchase_Date,Product_Category,Price_Paid,Country,\n"
        "TXN001, john doe,JOHN@EXAMPLE.COM,29,12-03-2026,electronics,1200,USA,\n"
        "TXN002,mike brown,mike@brown,twenty,Mar-19,2026,Home,$45.50,U.K.\n"
        "TXN002,mike brown,mike@brown,twenty,Mar-19,2026,Home,$45.50,U.K.\n"
        "TXN003,,valid@example.com,150,20-03-2026,,N/A,United States,\n"
    ).encode()
    response = client.post(
        "/api/clean",
        files={"file": ("messy transactions.csv", io.BytesIO(payload), "text/csv")},
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-disposition"] == (
        'attachment; filename="messy_transactions-sofidr-cleaned.csv"'
    )
    assert response.headers["x-sofidr-duplicates-removed"] == "1"
    assert response.headers["x-sofidr-completeness-after"] == "100.0"
    assert "repaired_2_malformed_rows" in response.headers["x-sofidr-cleaning-steps"]

    cleaned = pd.read_csv(io.BytesIO(response.content))
    assert len(cleaned) == 3
    assert cleaned.isna().sum().sum() == 0
    assert cleaned.loc[0, "Customer_Name"] == "John Doe"
    assert cleaned.loc[0, "Email"] == "john@example.com"
    assert cleaned.loc[0, "Country"] == "United States"
    assert cleaned.loc[1, "Purchase_Date"] == "2026-03-19"
    assert cleaned.loc[1, "Price_Paid"] == 45.5
    assert cleaned.loc[1, "Country"] == "United Kingdom"
    assert cleaned["Age"].between(0, 120).all()
